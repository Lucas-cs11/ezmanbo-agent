"""
output_bom.py — 专业工程级 BOM 生成模块

参照：《电子元器件行业 BOM 专业模板 v1.0》
结构层次：
  0. 文档控制信息（16 字段）
  1. BOM 主表（29 列，符合 EBOM/PCBA 标准）
  2. AVL/AML 替代料批准表
  3. 供应链风险子表
  4. 合规与质量证据表
  5. 变更记录
  6. 发布前检查清单

Excel 导出：4 个 Sheet（BOM主表 / 供应链风险 / AVL替代料 / 合规证据）
"""

from __future__ import annotations

import io
from typing import List, Optional, Tuple
from datetime import datetime, timezone

from .schemas import SelectionReport, ScoredPart, PartIR


# ──────────────────────────────────────────────────────────────────────────────
# 基础工具
# ──────────────────────────────────────────────────────────────────────────────

def _now_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _today() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def _lifecycle_normalized(status: Optional[str]) -> str:
    """标准化为受控词表枚举值。"""
    if not status:
        return "Unknown"
    s = status.strip().lower()
    if s in ("active", "production", "in production", "生产中", "量产"):
        return "Active"
    if s in ("nrnd", "nrfnd", "not recommended for new designs", "不推荐新设计"):
        return "NRND"
    if s in ("ltb", "last time buy", "最后一次采购"):
        return "LTB"
    if s in ("eol", "end of life", "停产", "生命周期末"):
        return "EOL"
    if s in ("obsolete", "discontinued", "已停产"):
        return "Obsolete"
    return "Unknown"


def _lifecycle_cn(status: Optional[str]) -> str:
    """带中文说明的生命周期标签。"""
    mapping = {
        "Active": "Active / 量产",
        "NRND": "NRND / 不推荐新设计",
        "LTB": "LTB / 最后采购",
        "EOL": "EOL / 停产中",
        "Obsolete": "Obsolete / 已停产",
        "Unknown": "Unknown / 待确认",
    }
    return mapping.get(_lifecycle_normalized(status), status or "Unknown / 待确认")


def _infer_mount_status(pkg: Optional[str]) -> str:
    """从封装代码推断贴装方式。"""
    if not pkg:
        return "SMT"
    p = pkg.upper()
    if any(t in p for t in ("DIP", "PDIP", "THT", "SIP", "RADIAL", "AXIAL", "TO220", "TO-220", "TO3", "TO-3")):
        return "TH"
    return "SMT"


def _infer_side(pkg: Optional[str]) -> str:
    return "Top"  # 默认顶层；实际由 PCB 设计师确定


def _infer_category(part: PartIR) -> str:
    """从 PartIR 推断器件分类。"""
    cat = (getattr(part, "category", None) or "").strip()
    if cat:
        # 规范化分类名称
        cl = cat.lower()
        if any(k in cl for k in ("buck", "boost", "dc-dc", "dcdc", "converter", "switching")):
            return "IC / DC-DC Converter"
        if any(k in cl for k in ("ldo", "linear reg")):
            return "IC / LDO Regulator"
        if any(k in cl for k in ("pmic",)):
            return "IC / PMIC"
        if any(k in cl for k in ("mcu", "microcontroller")):
            return "IC / MCU"
        if any(k in cl for k in ("capacitor", "电容")):
            return "Capacitor"
        if any(k in cl for k in ("resistor", "电阻")):
            return "Resistor"
        if any(k in cl for k in ("inductor", "电感")):
            return "Inductor"
        if any(k in cl for k in ("mosfet", "fet")):
            return "MOSFET"
        if any(k in cl for k in ("diode", "二极管")):
            return "Diode"
        return cat
    # 从 topology 推断
    topo = (getattr(part, "topology", None) or "").lower()
    if "buck" in topo:
        return "IC / Buck Converter"
    if "boost" in topo:
        return "IC / Boost Converter"
    if "ldo" in topo:
        return "IC / LDO Regulator"
    return "IC"


def _infer_supply_risk(part: PartIR) -> str:
    """从库存、生命周期推断供应链风险等级。"""
    lc = _lifecycle_normalized(part.lifecycle_status)
    stock = part.stock

    if lc in ("Obsolete", "EOL"):
        return "High"
    if lc == "LTB":
        return "High"
    if lc == "NRND":
        return "Medium"

    if stock is None:
        return "Medium"
    if stock == 0:
        return "High"
    if stock < 100:
        return "High"
    if stock < 500:
        return "Medium"

    return "Low"


def _is_critical(part: PartIR, s: ScoredPart) -> str:
    """判断是否为关键件（Critical Part）。"""
    # 车规认证 → 关键件
    if part.automotive_grade:
        return "Yes"
    # 主推荐器件 → 关键件
    if s.recommendation_level == "recommended":
        return "Yes"
    # 生命周期异常 → 关键件
    lc = _lifecycle_normalized(part.lifecycle_status)
    if lc in ("NRND", "LTB", "EOL", "Obsolete"):
        return "Yes"
    return "No"


def _infer_rating(part: PartIR) -> str:
    """生成额定值描述字符串。"""
    parts = []
    if part.input_voltage_max_v is not None:
        parts.append(f"Vin≤{part.input_voltage_max_v:.0f}V")
    if part.output_voltage_v is not None:
        parts.append(f"Vout={part.output_voltage_v}V")
    if part.output_current_max_a is not None:
        parts.append(f"Iout≤{part.output_current_max_a}A")
    temp_parts = []
    if part.temperature_min_c is not None:
        temp_parts.append(str(int(part.temperature_min_c)))
    if part.temperature_max_c is not None:
        temp_parts.append(str(int(part.temperature_max_c)))
    if temp_parts:
        parts.append(f"T={'/'.join(temp_parts)}°C")
    return "; ".join(parts) if parts else "N/A"


def _infer_value(part: PartIR) -> str:
    """生成 Value 字段（器件核心规格值）。"""
    if part.output_voltage_v is not None:
        return f"{part.output_voltage_v}V"
    topo = (getattr(part, "topology", None) or "").upper()
    if topo:
        return topo
    return part.part_number or "N/A"


def _description_short(part: PartIR) -> str:
    """生成规格描述（Description 字段）。"""
    cat = _infer_category(part)
    rating = _infer_rating(part)
    pkg = part.package or ""
    grade = "AEC-Q100" if part.automotive_grade else ""
    dom = "国产" if getattr(part, "is_domestic", False) else ""
    tags = " ".join(filter(None, [grade, dom]))
    desc_parts = [cat]
    if rating != "N/A":
        desc_parts.append(rating)
    if pkg:
        desc_parts.append(pkg)
    if tags:
        desc_parts.append(f"[{tags}]")
    return " | ".join(desc_parts)


def _rohs_status(part: PartIR) -> str:
    """RoHS 合规状态推断。"""
    # 车规认证器件通常 RoHS 合规
    if part.automotive_grade:
        return "Compliant"
    src = (part.source or "").lower()
    if src in ("ezplm", "api", "ezplm_api"):
        return "Compliant"  # eZ-PLM 平台数据通常符合 RoHS
    return "Unknown"


def _reach_status(part: PartIR) -> str:
    if part.automotive_grade:
        return "Compliant"
    return "Unknown"


def _alternate_mpns(part: PartIR, all_parts: List[ScoredPart]) -> str:
    """生成已批准替代料字符串（来自 replacement_for + 备选清单）。"""
    alts = []
    # replacement_for 列表（该器件可替代哪些料号）
    for ref in getattr(part, "replacement_for", []):
        if ref and ref not in alts:
            alts.append(ref)
    # 其他备选器件的 MPN（最多取 2 条）
    backup_parts = [
        s.part.part_number for s in all_parts
        if s.recommendation_level == "backup"
        and s.part.part_number != part.part_number
    ]
    for bp in backup_parts[:2]:
        if bp not in alts:
            alts.append(bp)
    return "; ".join(alts) if alts else "N/A"


# ──────────────────────────────────────────────────────────────────────────────
# 主 Markdown BOM 生成函数
# ──────────────────────────────────────────────────────────────────────────────

def generate_bom(report: SelectionReport, rag_context: str = "") -> str:
    """
    生成专业工程级 EBOM（Markdown）。

    符合《电子元器件行业 BOM 专业模板 v1.0》规范，包含：
    - 文档控制信息（16 字段）
    - 29 列 BOM 主表
    - AVL/AML 替代料批准表
    - 供应链风险子表
    - 合规与质量证据表
    - 变更记录
    - 发布前检查清单
    """
    c = report.constraints
    scored = report.candidates or []
    rec = [s for s in scored if s.recommendation_level == "recommended"]
    bak = [s for s in scored if s.recommendation_level == "backup"]
    all_bom = rec + bak

    bom_id = f"BOM-{report.request_id[:8].upper()}"
    today = _today()
    grade = c.grade or "Industrial"

    # ── 推断产品/拓扑信息 ─────────────────────────────────────────
    topo = c.topology or "未指定"
    app = getattr(c, "application", None) or "电源转换模块"
    vin_range = (
        f"{c.input_voltage_min_v}–{c.input_voltage_max_v}V"
        if (c.input_voltage_min_v is not None and c.input_voltage_max_v is not None)
        else (f"{c.input_voltage_nominal_v}V 标称" if c.input_voltage_nominal_v else "未指定")
    )
    vout_str = f"{c.output_voltage_v}V" if c.output_voltage_v else "未指定"
    iout_str = f"{c.output_current_a}A" if c.output_current_a else "未指定"
    temp_str = (
        f"{c.temperature_min_c}~{c.temperature_max_c}°C"
        if (c.temperature_min_c is not None and c.temperature_max_c is not None)
        else "未指定"
    )

    lines: List[str] = []

    # ════════════════════════════════════════════════════════════
    # 文件标题
    # ════════════════════════════════════════════════════════════
    lines += [
        f"# 工程物料清单 (Engineering BOM)",
        "",
        f"> **{app}** · {topo} · {grade} 级",
        f"> 原始需求：{c.raw_input}",
        "",
        "---",
        "",
    ]

    # ════════════════════════════════════════════════════════════
    # 第一节：文档控制信息
    # ════════════════════════════════════════════════════════════
    lines += [
        "## 1. 文档控制信息",
        "",
        "| 字段 | 内容 |",
        "|---|---|",
        f"| 项目名称 | {app} |",
        f"| 产品/模块名称 | {topo} 电源模块 |",
        f"| BOM 编号 | `{bom_id}` |",
        f"| BOM 版本 | Rev A（初始发布）|",
        f"| 适用硬件版本 | TBD（由硬件工程师填写）|",
        f"| 适用配置/变体 | Default |",
        f"| 基准数量 | 1 set / 1 PCBA |",
        f"| 输出日期 | {today} |",
        f"| 编制人 | eZ-PLM Agent（自动生成）|",
        f"| 审核人 | ＜硬件工程师 / 采购＞ |",
        f"| 批准人 | ＜项目负责人＞ |",
        f"| 状态 | **Draft**（需人工审核后发布）|",
        f"| 关联 ECO/ECN | ECO-{report.request_id[:6].upper()}-001 |",
        f"| 设计工具与源文件 | TBD |",
        f"| 应用等级 | {grade} |",
        f"| 数据来源 | eZ-PLM 平台 API（实时查询）|",
        "",
        "---",
        "",
    ]

    # ════════════════════════════════════════════════════════════
    # 第二节：BOM 主表（29 列 EBOM）
    # ════════════════════════════════════════════════════════════
    lines += [
        "## 2. BOM 主表（EBOM）",
        "",
        "> **说明**：RefDes（参考位号）需在 PCB 布局完成后由硬件工程师补充。",
        "> Footprint、MSL 等字段需从数据手册或制造商官网核实后填写。",
        "> 未标注具体值的字段以 `TBD` 或 `Unknown` 表示，发布前须补全（见第 7 节检查清单）。",
        "",
    ]

    # 29 列主表表头
    bom_header = (
        "| Item No. | Level | Parent Assembly | RefDes | Qty | UOM "
        "| Mount | Side | Category | Description "
        "| Value | Tolerance | Rating "
        "| Package | Footprint "
        "| Manufacturer | MPN "
        "| Lifecycle | Primary Supplier | Supplier PN "
        "| Alternate MPNs | RoHS | REACH | MSL "
        "| Critical | Supply Risk | DNP | Variant | Notes |"
    )
    bom_sep = (
        "|---:|---:|---|---|---:|---|"
        "---|---|---|---"
        "|---|---|---"
        "|---|---"
        "|---|---"
        "|---|---|---"
        "|---|---|---|---"
        "|---|---|---|---|---|"
    )
    lines += [bom_header, bom_sep]

    item_no = 10
    for s in all_bom:
        p = s.part
        category = _infer_category(p)
        description = _description_short(p)
        value = _infer_value(p)
        rating = _infer_rating(p)
        mount = _infer_mount_status(p.package)
        side = _infer_side(p.package)
        lc = _lifecycle_normalized(p.lifecycle_status)
        rohs = _rohs_status(p)
        reach = _reach_status(p)
        supply_risk = _infer_supply_risk(p)
        critical = _is_critical(p, s)
        alt_mpns = _alternate_mpns(p, all_bom)
        mfr = p.manufacturer or "TBD"
        pkg = p.package or "TBD"
        dom = " 🇨🇳" if getattr(p, "is_domestic", False) else ""

        # 风险标记
        lc_icon = "⚠ " if lc in ("NRND", "LTB", "EOL", "Obsolete") else ""
        risk_icon = {"High": "🔴", "Medium": "🟡", "Low": "✅"}.get(supply_risk, "")

        row = (
            f"| {item_no} | 2 | PCBA Assembly "
            f"| TBD | 1 | pcs "
            f"| {mount} | {side} | {category} | {description[:48]} "
            f"| {value} | N/A | {rating[:30]} "
            f"| {pkg} | TBD "
            f"| {mfr}{dom} | `{p.part_number}` "
            f"| {lc_icon}{lc} | eZ-PLM | TBD "
            f"| {alt_mpns[:40]} | {rohs} | {reach} | TBD "
            f"| {critical} | {risk_icon} {supply_risk} | No | Default | — |"
        )
        lines.append(row)
        item_no += 10

    if not all_bom:
        lines.append("| — | — | — | — | — | — | — | — | — | 无候选器件 | — | — | — | — | — | — | — | — | — | — | — | — | — | — | — | — | — | — | — |")

    lines += ["", "---", ""]

    # ════════════════════════════════════════════════════════════
    # 第三节：AVL/AML 替代料批准表
    # ════════════════════════════════════════════════════════════
    lines += [
        "## 3. AVL/AML 替代料批准表",
        "",
        "> 替代料须经工程/采购/质量或客户批准后方可在 BOM 中使用。",
        "> Form-Fit-Function 等级：FFF = 完全等价，CFR = 条件替代，需额外验证。",
        "",
        "| AVL ID | Item No. | 原始 MPN | 替代制造商 | 替代 MPN | FFF 等级 | 适用范围 | 限制条件 | 批准人 | 批准日期 |",
        "|---|---:|---|---|---|---|---|---|---|---|",
    ]

    avl_id = 1
    # 推荐器件 → 备选器件形成替代关系
    if rec and bak:
        for backup_s in bak[:5]:
            bp = backup_s.part
            for rec_s in rec[:1]:
                rp = rec_s.part
                lines.append(
                    f"| AVL-{avl_id:03d} | 10 | `{rp.part_number}` "
                    f"| {bp.manufacturer or 'TBD'} | `{bp.part_number}` "
                    f"| CFR | Default 配置 "
                    f"| 需验证关键参数；建议样品测试 "
                    f"| ＜待批准＞ | TBD |"
                )
                avl_id += 1

    # 器件自身的 replacement_for 记录
    for s in all_bom:
        p = s.part
        for ref_mpn in getattr(p, "replacement_for", [])[:2]:
            lines.append(
                f"| AVL-{avl_id:03d} | — | `{ref_mpn}` "
                f"| {p.manufacturer or 'TBD'} | `{p.part_number}` "
                f"| CFR | 原料停产替代 "
                f"| 需评估引脚/封装兼容性 "
                f"| ＜待批准＞ | TBD |"
            )
            avl_id += 1

    if avl_id == 1:
        lines.append("| — | — | 暂无替代料记录；量产前须补充至少一条替代料 | — | — | — | — | — | — | — |")

    lines += ["", "---", ""]

    # ════════════════════════════════════════════════════════════
    # 第四节：供应链风险表
    # ════════════════════════════════════════════════════════════
    lines += [
        "## 4. 供应链风险表",
        "",
        "| Item No. | MPN | Lifecycle | Sources | Lead Time | 当前库存 | Supply Risk | 风险原因 | 缓解计划 | 负责人 | 截止日期 |",
        "|---:|---|---|---:|---|---:|---|---|---|---|---|",
    ]

    item_no = 10
    for s in all_bom:
        p = s.part
        lc = _lifecycle_normalized(p.lifecycle_status)
        risk = _infer_supply_risk(p)
        stock_str = str(p.stock) if p.stock is not None else "未知"
        risk_icon = {"High": "🔴 High", "Medium": "🟡 Medium", "Low": "✅ Low"}.get(risk, risk)

        # 风险原因
        reasons = []
        if lc in ("EOL", "Obsolete"):
            reasons.append("已停产")
        elif lc == "NRND":
            reasons.append("不推荐新设计")
        elif lc == "LTB":
            reasons.append("最后采购期限临近")
        if p.stock is not None and p.stock < 100:
            reasons.append(f"库存不足（{p.stock} 件）")
        if not reasons:
            reasons.append("暂无明显风险信号")

        # 缓解计划
        if lc in ("EOL", "Obsolete"):
            mitigation = "启动替代料验证；评估 LTB 数量"
        elif lc == "NRND":
            mitigation = "预研替代料；监控 PCN/PDN 通知"
        elif p.stock is not None and p.stock < 100:
            mitigation = "立即预购；建立安全库存"
        else:
            mitigation = "纳入常规 PCN/EOL 监控"

        lines.append(
            f"| {item_no} | `{p.part_number}` | {lc} | 1 | TBD | {stock_str} "
            f"| {risk_icon} | {'; '.join(reasons)} | {mitigation} | ＜待指定＞ | TBD |"
        )
        item_no += 10

    if not all_bom:
        lines.append("| — | 无候选器件 | — | — | — | — | — | — | — | — | — |")

    lines += ["", "---", ""]

    # ════════════════════════════════════════════════════════════
    # 第五节：合规与质量证据表
    # ════════════════════════════════════════════════════════════
    lines += [
        "## 5. 合规与质量证据表",
        "",
        "| Item No. | MPN | RoHS 证据 | REACH 证据 | 安规/UL 证据 | 可靠性证据 | 追溯要求 | 证据负责人 | 复核日期 | 状态 |",
        "|---:|---|---|---|---|---|---|---|---|---|",
    ]

    item_no = 10
    for s in all_bom:
        p = s.part
        rohs_ev = "eZ-PLM 平台 / 待制造商正式声明（E1→E3）"
        reach_ev = "待制造商 REACH/SVHC 声明"
        safety_ev = "N/A（DC-DC/LDO 无直接 UL 要求）" if "IC" in _infer_category(p) else "TBD"
        rel_ev = "AEC-Q100 测试报告" if p.automotive_grade else "Datasheet（E2）"
        trace = "Lot"
        status = "Pending" if not p.datasheet_url else "Partial"

        lines.append(
            f"| {item_no} | `{p.part_number}` "
            f"| {rohs_ev} | {reach_ev} "
            f"| {safety_ev} | {rel_ev} "
            f"| {trace} | ＜待指定＞ | TBD | {status} |"
        )
        item_no += 10

    if not all_bom:
        lines.append("| — | 无候选器件 | — | — | — | — | — | — | — | — |")

    lines += ["", "---", ""]

    # ════════════════════════════════════════════════════════════
    # 第六节：变更记录
    # ════════════════════════════════════════════════════════════
    lines += [
        "## 6. 变更记录",
        "",
        "| BOM Rev | 日期 | ECO/ECN | 变更类型 | 变更内容 | 受影响 Item | 编制 | 审核 | 批准 |",
        "|---|---|---|---|---|---|---|---|---|",
        f"| A | {today} | ECO-{report.request_id[:6].upper()}-001 "
        f"| Initial Release | eZ-PLM Agent 自动生成初稿，需人工审核后正式发布 "
        f"| All | eZ-PLM Agent | ＜待填写＞ | ＜待填写＞ |",
        "",
        "---",
        "",
    ]

    # ════════════════════════════════════════════════════════════
    # 第七节：发布前检查清单
    # ════════════════════════════════════════════════════════════
    lines += [
        "## 7. 发布前检查清单",
        "",
        "### 7.1 设计一致性",
        "",
    ]

    # 动态检测字段完整性
    missing_refdes = True  # RefDes 始终由工程师补充
    missing_footprint = all(not s.part.package for s in all_bom)
    missing_lifecycle = any(not s.part.lifecycle_status for s in all_bom)

    checks_design = [
        ("- [ ]", "BOM 中所有 RefDes 与原理图一致（需人工填写）"),
        ("- [ ]", "BOM 中所有 Footprint 与 PCB Layout 一致（需人工填写）"),
        ("- [ ]", "Qty Per 与参考位号数量一致"),
        ("- [ ]", "DNP 器件已明确标识，且与装配图/坐标文件一致"),
        ("- [x]" if not missing_lifecycle else "- [ ]",
         "所有器件生命周期状态已确认（" +
         ("✅ 已确认" if not missing_lifecycle else f"⚠ {sum(1 for s in all_bom if not s.part.lifecycle_status)} 条待确认") + "）"),
    ]
    for mark, text in checks_design:
        lines.append(f"{mark} {text}")

    lines += [
        "",
        "### 7.2 采购与供应链",
        "",
    ]

    high_risk_parts = [s for s in all_bom if _infer_supply_risk(s.part) == "High"]
    no_mfr = [s for s in all_bom if not s.part.manufacturer]
    eol_parts = [s for s in all_bom if _lifecycle_normalized(s.part.lifecycle_status) in ("EOL", "Obsolete")]
    avl_count = len(bak)

    checks_supply = [
        ("- [x]" if not no_mfr else "- [ ]",
         f"所有采购件均有 Manufacturer 和完整 MPN（" +
         ("✅ 完整" if not no_mfr else f"⚠ {len(no_mfr)} 条制造商信息缺失") + "）"),
        ("- [ ]", "所有采购件均有首选供应商和供应商料号（TBD，需采购填写）"),
        ("- [x]" if not high_risk_parts else "- [ ]",
         f"单一来源或长交期物料已标记风险（" +
         ("✅ 无高风险" if not high_risk_parts else
          f"⚠ {len(high_risk_parts)} 条高供应风险：{', '.join(s.part.part_number for s in high_risk_parts[:3])}") + "）"),
        ("- [x]" if avl_count > 0 else "- [ ]",
         f"关键件或高风险物料已有批准替代料（" +
         (f"✅ {avl_count} 条备选" if avl_count > 0 else "⚠ 无备选器件，量产前须补充 AVL") + "）"),
        ("- [ ]" if eol_parts else "- [x]",
         f"EOL/Obsolete 物料已处理（" +
         ("✅ 无停产器件" if not eol_parts else
          f"⚠ {len(eol_parts)} 条停产：{', '.join(s.part.part_number for s in eol_parts[:3])}") + "）"),
        ("- [ ]", "MOQ、MPQ、价格、交期已注明更新时间"),
    ]
    for mark, text in checks_supply:
        lines.append(f"{mark} {text}")

    lines += [
        "",
        "### 7.3 制造与质量",
        "",
        "- [ ] MSL、回流温度、包装方式等制造要求已记录",
        "- [ ] 关键件、安规件、客户指定料已标识",
        "- [ ] 检验要求、COC/COA、批次追溯要求已明确",
        "- [ ] BOM 与坐标文件、装配图、Gerber 资料包版本一致",
        "",
        "### 7.4 合规",
        "",
    ]

    no_datasheet = [s for s in all_bom if not s.part.datasheet_url]
    checks_compliance = [
        ("- [ ]", "RoHS 状态已由制造商正式声明确认（E3 级证据）"),
        ("- [ ]", "REACH/SVHC 状态已确认，或明确不适用"),
        ("- [x]" if not no_datasheet else "- [ ]",
         f"数据手册链接已归档（" +
         ("✅ 完整" if not no_datasheet else f"⚠ {len(no_datasheet)} 条缺失数据手册") + "）"),
        ("- [ ]", "合规证据有文件编号、负责人和复核日期"),
    ]
    for mark, text in checks_compliance:
        lines.append(f"{mark} {text}")

    lines += [
        "",
        "### 7.5 变更与发布",
        "",
        "- [ ] BOM 版本、硬件版本、ECO/ECN 编号完整",
        "- [ ] 变更前后差异已审查",
        "- [ ] 研发、采购、制造、质量、项目管理均已批准",
        "- [ ] 发布状态更新为 `Released` 后禁止未授权手工修改",
        "",
        "---",
        "",
    ]

    # ════════════════════════════════════════════════════════════
    # 第八节：选型评分概览（保留原有评分信息）
    # ════════════════════════════════════════════════════════════
    lines += [
        "## 8. 选型评分概览",
        "",
        "| # | MPN | 制造商 | 评分 | 参数匹配 | 供应链 | 国产化 | 推荐等级 | 风险等级 |",
        "|---:|---|---|---:|---:|---:|---:|---|---|",
    ]

    risk_per_part = {}
    if report.risks and report.risks.per_part_risks:
        for r in report.risks.per_part_risks:
            risk_per_part[r.get("part_number", "")] = r

    for idx, s in enumerate(all_bom, 1):
        p = s.part
        sc = s.score
        part_risk = risk_per_part.get(p.part_number)
        lc_code = part_risk["risk_level_code"] if part_risk else "—"
        lc_emoji = {"L1": "✅", "L2": "🟡", "L3": "🟠", "L4": "🔴", "L5": "⛔"}.get(lc_code, "?")
        dom_tag = " 🇨🇳" if getattr(p, "is_domestic", False) else ""
        lines.append(
            f"| {idx} | `{p.part_number}` | {p.manufacturer or '—'}{dom_tag} "
            f"| **{sc.total_score:.1f}** | {sc.parameter_match_score:.0f} "
            f"| {sc.supply_risk_score:.0f} | {sc.domestic_score:.0f} "
            f"| {s.recommendation_level or '—'} | {lc_emoji} {lc_code} |"
        )

    lines += ["", "---", ""]

    # ════════════════════════════════════════════════════════════
    # RAG 工程知识参考（若有）
    # ════════════════════════════════════════════════════════════
    if rag_context:
        lines += [
            "## 9. 工程知识库参考",
            "",
            rag_context,
            "",
            "---",
            "",
        ]

    # ════════════════════════════════════════════════════════════
    # 页脚免责声明
    # ════════════════════════════════════════════════════════════
    lines += [
        f"*本文档由 eZ-PLM Agent 自动生成，生成时间：{_now_str()}。*",
        "*当前状态：**Draft**，须经硬件研发、采购、制造、质量、项目管理人员审核批准后方可发布为 Released。*",
        "*RefDes、Footprint、Supplier PN、MOQ、交期、MSL 等标注 TBD 字段须由相关人员补全。*",
        "*最终选型须结合完整数据手册、实际工况、企业 AVL/AML 及 DFM/DFT 审查综合评估。*",
        "",
        f"**文件命名建议**：`{bom_id.replace('BOM-', '')}_{topo.replace(' ', '_')}_BOM_HW-TBD_REV-A_Default_{today.replace('-', '')}.md`",
    ]

    return "\n".join(lines)


# ──────────────────────────────────────────────────────────────────────────────
# Excel 导出（4 Sheet 专业版）
# ──────────────────────────────────────────────────────────────────────────────

def generate_bom_excel(report: SelectionReport) -> bytes:
    """
    生成专业工程级 BOM Excel 文件（4 个 Sheet）。

    Sheet 1 — BOM 主表（29 列 EBOM）
    Sheet 2 — 供应链风险表
    Sheet 3 — AVL/AML 替代料批准表
    Sheet 4 — 合规与质量证据表
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return b""

    wb = openpyxl.Workbook()

    # ── 全局样式 ──────────────────────────────────────────────
    FONT_BODY = Font(name="Calibri", size=9)
    FONT_HEADER = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    FONT_TITLE = Font(name="Calibri", size=12, bold=True, color="1F3864")
    FONT_META = Font(name="Calibri", size=9, bold=True)

    FILL_HEADER = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    FILL_REC = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    FILL_BAK = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    FILL_HIGH = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    FILL_MED = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    FILL_LOW = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    FILL_ALT = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    BORDER_THIN = Border(
        left=Side(style="thin", color="BDBDBD"),
        right=Side(style="thin", color="BDBDBD"),
        top=Side(style="thin", color="BDBDBD"),
        bottom=Side(style="thin", color="BDBDBD"),
    )

    c = report.constraints
    scored = report.candidates or []
    rec_parts = [s for s in scored if s.recommendation_level == "recommended"]
    bak_parts = [s for s in scored if s.recommendation_level == "backup"]
    all_bom = rec_parts + bak_parts

    today = _today()
    bom_id = f"BOM-{report.request_id[:8].upper()}"

    # ────────────────────────────────────────────────────────────
    # Sheet 1: BOM 主表
    # ────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "BOM主表"

    # 标题
    ws1.merge_cells("A1:AC1")
    t = ws1["A1"]
    t.value = f"Engineering BOM — {c.raw_input[:80]}"
    t.font = FONT_TITLE
    t.alignment = ALIGN_LEFT
    ws1.row_dimensions[1].height = 28

    # 元数据（2 行）
    meta_r1 = [("BOM 编号", bom_id), ("版本", "Rev A"), ("日期", today), ("状态", "Draft")]
    meta_r2 = [
        ("拓扑", c.topology or "—"),
        ("Vin", f"{c.input_voltage_nominal_v}V" if c.input_voltage_nominal_v else "—"),
        ("Vout", f"{c.output_voltage_v}V" if c.output_voltage_v else "—"),
        ("Iout", f"{c.output_current_a}A" if c.output_current_a else "—"),
    ]
    col = 1
    for k, v in meta_r1:
        ws1.cell(2, col, k).font = FONT_META
        ws1.cell(2, col + 1, v).font = FONT_BODY
        col += 3
    col = 1
    for k, v in meta_r2:
        ws1.cell(3, col, k).font = FONT_META
        ws1.cell(3, col + 1, v).font = FONT_BODY
        col += 3
    ws1.row_dimensions[2].height = 16
    ws1.row_dimensions[3].height = 16

    # 29 列表头（第 5 行）
    HEADERS_BOM = [
        "Item No.", "Level", "Parent Assembly", "RefDes", "Qty", "UOM",
        "Mount", "Side", "Category", "Description",
        "Value", "Tolerance", "Rating",
        "Package", "Footprint",
        "Manufacturer", "MPN",
        "Lifecycle", "Primary Supplier", "Supplier PN",
        "Alt MPNs", "RoHS", "REACH", "MSL",
        "Critical", "Supply Risk", "DNP", "Variant", "Notes",
    ]
    HDR_ROW = 5
    for ci, h in enumerate(HEADERS_BOM, 1):
        cell = ws1.cell(HDR_ROW, ci, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = BORDER_THIN
    ws1.row_dimensions[HDR_ROW].height = 28

    # 数据行
    item_no = 10
    for data_idx, s in enumerate(all_bom, 1):
        p = s.part
        row_idx = HDR_ROW + data_idx
        category = _infer_category(p)
        mount = _infer_mount_status(p.package)
        lc = _lifecycle_normalized(p.lifecycle_status)
        rohs = _rohs_status(p)
        reach = _reach_status(p)
        supply_risk = _infer_supply_risk(p)
        critical = _is_critical(p, s)
        alt_mpns = _alternate_mpns(p, all_bom)
        description = _description_short(p)
        rating = _infer_rating(p)
        value = _infer_value(p)

        row_data = [
            item_no, 2, "PCBA Assembly", "TBD", 1, "pcs",
            mount, "Top", category, description,
            value, "N/A", rating,
            p.package or "TBD", "TBD",
            p.manufacturer or "TBD", p.part_number,
            lc, "eZ-PLM", "TBD",
            alt_mpns, rohs, reach, "TBD",
            critical, supply_risk, "No", "Default", "",
        ]

        # 背景色：推荐绿，备选黄
        fill_row = FILL_REC if s.recommendation_level == "recommended" else (
            FILL_BAK if s.recommendation_level == "backup" else FILL_ALT
        )

        for ci, val in enumerate(row_data, 1):
            cell = ws1.cell(row_idx, ci, val)
            cell.font = FONT_BODY
            cell.border = BORDER_THIN
            cell.fill = fill_row
            cell.alignment = ALIGN_CENTER if ci in (1, 2, 5, 6, 7, 8, 22, 23, 24, 25, 26, 27) else ALIGN_LEFT

            # 高风险着色
            if ci == 26 and supply_risk == "High":
                cell.fill = FILL_HIGH
            elif ci == 18 and lc in ("EOL", "Obsolete"):
                cell.fill = FILL_HIGH
            elif ci == 18 and lc in ("NRND", "LTB"):
                cell.fill = FILL_MED

        ws1.row_dimensions[row_idx].height = 18
        item_no += 10

    # 列宽设置（29 列）
    col_widths = [
        7, 6, 16, 12, 5, 5,
        8, 6, 18, 40,
        10, 9, 28,
        12, 14,
        16, 22,
        12, 16, 16,
        30, 10, 10, 8,
        9, 12, 6, 10, 16,
    ]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    ws1.freeze_panes = f"A{HDR_ROW + 1}"
    ws1.sheet_view.showGridLines = True

    # ────────────────────────────────────────────────────────────
    # Sheet 2: 供应链风险表
    # ────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("供应链风险")
    ws2.merge_cells("A1:K1")
    ws2["A1"].value = f"供应链风险表 — {bom_id}"
    ws2["A1"].font = FONT_TITLE
    ws2.row_dimensions[1].height = 24

    risk_headers = [
        "Item No.", "MPN", "制造商", "Lifecycle",
        "库存", "Supply Risk", "风险原因", "缓解计划",
        "负责人", "截止日期", "状态",
    ]
    for ci, h in enumerate(risk_headers, 1):
        cell = ws2.cell(3, ci, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = BORDER_THIN
    ws2.row_dimensions[3].height = 22

    item_no = 10
    for data_idx, s in enumerate(all_bom, 1):
        p = s.part
        row_idx = 3 + data_idx
        lc = _lifecycle_normalized(p.lifecycle_status)
        risk = _infer_supply_risk(p)

        reasons = []
        if lc in ("EOL", "Obsolete"):
            reasons.append("已停产")
        elif lc in ("NRND", "LTB"):
            reasons.append(f"生命周期风险（{lc}）")
        if p.stock is not None and p.stock < 100:
            reasons.append(f"库存低（{p.stock}件）")
        if not reasons:
            reasons.append("暂无明显风险")

        if lc in ("EOL", "Obsolete"):
            mitigation = "启动替代料验证；评估 LTB"
        elif lc == "NRND":
            mitigation = "预研替代料；订阅 PCN"
        elif p.stock is not None and p.stock < 100:
            mitigation = "立即预购；建立安全库存"
        else:
            mitigation = "常规 PCN/EOL 监控"

        row_data = [
            item_no,
            p.part_number,
            p.manufacturer or "TBD",
            lc,
            p.stock if p.stock is not None else "Unknown",
            risk,
            "; ".join(reasons),
            mitigation,
            "TBD",
            "TBD",
            "Open",
        ]

        fill_row = FILL_HIGH if risk == "High" else (FILL_MED if risk == "Medium" else FILL_LOW)

        for ci, val in enumerate(row_data, 1):
            cell = ws2.cell(row_idx, ci, val)
            cell.font = FONT_BODY
            cell.border = BORDER_THIN
            cell.fill = fill_row
            cell.alignment = ALIGN_CENTER if ci in (1, 4, 5, 6, 10, 11) else ALIGN_LEFT
        ws2.row_dimensions[row_idx].height = 18
        item_no += 10

    for ci, w in enumerate([7, 22, 14, 12, 8, 12, 36, 30, 12, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A4"

    # ────────────────────────────────────────────────────────────
    # Sheet 3: AVL/AML 替代料批准表
    # ────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("AVL替代料")
    ws3.merge_cells("A1:J1")
    ws3["A1"].value = f"AVL/AML 替代料批准表 — {bom_id}"
    ws3["A1"].font = FONT_TITLE
    ws3.row_dimensions[1].height = 24

    avl_headers = [
        "AVL ID", "Item No.", "原始 MPN", "替代制造商",
        "替代 MPN", "FFF 等级", "适用范围",
        "限制条件", "批准人", "批准日期",
    ]
    for ci, h in enumerate(avl_headers, 1):
        cell = ws3.cell(3, ci, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = BORDER_THIN
    ws3.row_dimensions[3].height = 22

    avl_id = 1
    avl_item_no = 10
    avl_row = 4

    if rec_parts and bak_parts:
        for backup_s in bak_parts[:5]:
            bp = backup_s.part
            for rec_s in rec_parts[:1]:
                rp = rec_s.part
                row_data = [
                    f"AVL-{avl_id:03d}", avl_item_no,
                    rp.part_number, bp.manufacturer or "TBD",
                    bp.part_number, "CFR", "Default 配置",
                    "需验证关键参数；建议样品测试",
                    "TBD", "TBD",
                ]
                for ci, val in enumerate(row_data, 1):
                    cell = ws3.cell(avl_row, ci, val)
                    cell.font = FONT_BODY
                    cell.border = BORDER_THIN
                    cell.fill = FILL_BAK
                    cell.alignment = ALIGN_CENTER if ci in (1, 2, 6, 9, 10) else ALIGN_LEFT
                ws3.row_dimensions[avl_row].height = 18
                avl_id += 1
                avl_row += 1
            avl_item_no += 10

    if avl_row == 4:
        ws3.cell(4, 1, "暂无替代料记录，量产前须补充").font = FONT_BODY

    for ci, w in enumerate([10, 8, 22, 16, 22, 10, 14, 36, 12, 12], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.freeze_panes = "A4"

    # ────────────────────────────────────────────────────────────
    # Sheet 4: 合规与质量证据
    # ────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("合规证据")
    ws4.merge_cells("A1:J1")
    ws4["A1"].value = f"合规与质量证据表 — {bom_id}"
    ws4["A1"].font = FONT_TITLE
    ws4.row_dimensions[1].height = 24

    comp_headers = [
        "Item No.", "MPN", "制造商", "RoHS 证据",
        "REACH 证据", "安规/UL 证据", "可靠性证据",
        "追溯要求", "证据负责人", "状态",
    ]
    for ci, h in enumerate(comp_headers, 1):
        cell = ws4.cell(3, ci, h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_HEADER
        cell.border = BORDER_THIN
    ws4.row_dimensions[3].height = 22

    item_no = 10
    for data_idx, s in enumerate(all_bom, 1):
        p = s.part
        row_idx = 3 + data_idx
        rohs_ev = "eZ-PLM 平台（E1）" if p.source in ("ezplm", "api") else "待补充"
        reach_ev = "待制造商声明（E3）"
        safety_ev = "N/A"
        rel_ev = "AEC-Q100（E4）" if p.automotive_grade else "Datasheet（E2）"
        status = "Partial" if p.datasheet_url else "Pending"

        row_data = [
            item_no, p.part_number, p.manufacturer or "TBD",
            rohs_ev, reach_ev, safety_ev, rel_ev,
            "Lot", "TBD", status,
        ]
        fill_row = FILL_LOW if status == "Partial" else FILL_MED
        for ci, val in enumerate(row_data, 1):
            cell = ws4.cell(row_idx, ci, val)
            cell.font = FONT_BODY
            cell.border = BORDER_THIN
            cell.fill = fill_row
            cell.alignment = ALIGN_CENTER if ci in (1, 8, 10) else ALIGN_LEFT
        ws4.row_dimensions[row_idx].height = 18
        item_no += 10

    for ci, w in enumerate([8, 22, 16, 22, 22, 18, 22, 12, 12, 10], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.freeze_panes = "A4"

    # ── 输出 ──────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
