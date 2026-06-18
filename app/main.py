import os
from dotenv import load_dotenv
load_dotenv()

from typing import Optional, Dict, Any, AsyncGenerator
from fastapi import FastAPI, Request, Body, UploadFile, File, Form
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse, StreamingResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import json
import time
import asyncio
import traceback
from .agent_orchestrator import analyze, replacement_report

# ── 会话级报告缓存（供 /report/{type} 端点使用）────────────────
# key=session_id, value=(SelectionReport, constraints)
_session_reports: Dict[str, Any] = {}
_session_constraints: Dict[str, Any] = {}
_DEFAULT_SESSION_ID = "__default__"  # 未提供 session_id 时的兜底key

app = FastAPI()

# ── CORS 配置，从环境变量读取白名单（M2）──────────────────────
_cors_origins_raw = os.environ.get("CORS_ORIGINS", "http://localhost:3000,http://localhost:3001").split(",")
_cors_origins = [o.strip() for o in _cors_origins_raw if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class AnalyzeRequest(BaseModel):
    user_input: str
    thinking_depth: str = "default"  # off | default | contemplation | exhaustive
    session_id: Optional[str] = None  # C1: 会话隔离

class ReplacementRequest(BaseModel):
    original_part_number: str

@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    """处理请求验证错误"""
    return JSONResponse(
        status_code=422,
        content={"detail": "请求体格式错误", "errors": exc.errors()},
    )

@app.get("/health")
async def health():
    return {"status": "ok"}

@app.post("/analyze")
async def analyze_endpoint(body: AnalyzeRequest):
    """Pipeline 模式：单次调用，返回完整 SelectionReport JSON。

    ── B4：语义缓存支持 ──
    - 响应头 X-Cache: HIT 表示命中缓存
    - 响应头 X-Cache: MISS 表示未命中缓存
    """
    try:
        session_id = body.session_id or _DEFAULT_SESSION_ID
        # ── B4：检查语义缓存 ──────────────────────────────────────
        from .semantic_cache import get_semantic_cache
        cache = get_semantic_cache()
        cache_result = cache.get(body.user_input)
        cache_header = "HIT" if cache_result is not None else "MISS"

        # P1: 优先使用 LangGraph 状态机（含 CriticNode + 自动放宽）
        try:
            from .langgraph_orchestrator import run_selection_pipeline
            result = run_selection_pipeline(body.user_input)
            if result.get("report"):
                report = result["report"]
            else:
                report = analyze(body.user_input)  # fallback
        except Exception:
            report = analyze(body.user_input)

        _session_reports[session_id] = report
        _session_constraints[session_id] = report.constraints
        return JSONResponse(
            content=report.dict(),
            headers={"X-Cache": cache_header}
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"detail": f"处理错误: {str(e)}"},
        )

@app.post("/replacement")
async def replacement_endpoint(body: ReplacementRequest):
    """替代器件查找。"""
    try:
        return replacement_report(body.original_part_number)
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"detail": f"处理错误: {str(e)}"},
        )


# ── Agent 端点 ──────────────────────────────────────────────────

class AgentRequest(BaseModel):
    user_input: str
    session_id: Optional[str] = None
    thinking_depth: str = "default"

@app.post("/agent/chat")
async def agent_chat_endpoint(body: AgentRequest):
    """ReAct Agent 模式：支持多轮对话，返回推理过程 + 工具调用记录。"""
    try:
        from .react_agent import run_agent
        result = run_agent(body.user_input, session_id=body.session_id,
                          thinking_depth=body.thinking_depth)

        # ── Memory：提取并记录用户称呼 ──────────────────────
        try:
            from .memory import update_user_name
            import re as _re
            name_m = _re.search(r'(?:我是|我叫|叫我)\s*[\u4e00-\u9fff\w]+', body.user_input)
            if name_m:
                name = _re.sub(r'(?:我是|我叫|叫我)\s*', '', name_m.group(0)).strip().rstrip('.,;!。，；！')
                if name and 1 < len(name) <= 10:
                    update_user_name(name)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"agent_chat: 记忆写入失败: {e}")

        return result
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"detail": f"Agent 错误: {str(e)}"},
        )

@app.get("/agent/sessions")
async def agent_sessions_endpoint():
    """列出当前活跃的 Agent 会话，含元数据（标题、消息数）。"""
    from langchain_core.messages import HumanMessage
    from .react_agent import _sessions
    sessions = []
    for sid, msgs in _sessions.items():
        # 从第一条用户消息提取标题
        title = "新的对话"
        for m in msgs:
            if isinstance(m, HumanMessage) and m.content:
                raw = str(m.content).strip()
                # 跳过内部注入的系统消息
                if raw.startswith("[选型上下文") or raw.startswith("[selection_context"):
                    continue
                title = raw[:30].replace("\n", " ")
                break
        sessions.append({
            "id": sid,
            "title": title,
            "message_count": len(msgs),
        })
    return {"sessions": sessions, "total": len(sessions)}


@app.post("/agent/init_session")
async def agent_init_session_endpoint(body: dict = Body(...)):
    """创建带预注入上下文的 Agent 会话（无 LLM 调用）。

    直接把选型摘要写入会话历史（AIMessage），
    后续 /agent/chat 调用时 Agent 可基于此上下文作答。

    请求体：
      - session_id: 会话 ID（可选，未提供时自动生成）
      - context: 上下文文本（与 payload 等效）
      - payload: 上下文文本（context 的别名）
      - context_type: 上下文类型标签（如 selection_context, general）
    """
    from .react_agent import get_or_create_session, _sessions
    from langchain_core.messages import AIMessage
    import os as _os

    session_id = body.get("session_id") or _os.urandom(8).hex()
    context = body.get("context") or body.get("payload", "").strip() if isinstance(body, dict) else ""
    ctx_type = body.get("context_type", "general") if isinstance(body, dict) else "general"

    if context:
        history = get_or_create_session(session_id)
        prefix = f"[{ctx_type}]" if ctx_type != "general" else "[选型上下文已同步]"
        history.append(AIMessage(content=f"{prefix}\n{context}"))
        _sessions[session_id] = history

    return {"session_id": session_id, "injected": bool(context), "context_type": ctx_type}


# ── 意图分类端点（三层分流机制）───────────────────────────────

class ClassifyRequest(BaseModel):
    user_input: str
    has_active_selection: bool = False
    accumulated_input: str = ""  # 跨轮累积的约束文本（分号分隔）
    thinking_depth: str = "default"

@app.post("/classify")
async def classify_endpoint(body: ClassifyRequest):
    """对用户输入进行意图分类。

    返回:
      {"intent": "selection"|"chat"|"adjustment"|"clarify",
       "reasoning": "LLM判断理由",
       "is_in_scope": true/false,
       ...}
    """
    from .intent_classifier import classify, extract_adjustment, get_last_classification
    from .constraint_checker import extract_constraints, check_completeness, build_clarification_response, merge_constraints

    intent = classify(body.user_input, body.has_active_selection,
                      accumulated_input=body.accumulated_input)
    result: dict = {"intent": intent}

    # 附加 LLM 分类详情（reasoning, device_category 等）
    llm_info = get_last_classification()
    if llm_info:
        result["reasoning"] = llm_info.get("reasoning", "")
        result["device_category"] = llm_info.get("device_category", "")

    if intent == "adjustment":
        result["adjustments"] = extract_adjustment(body.user_input)

    if intent == "clarify":
        # 如果有累积的前轮约束文本，合并后重新检查完整性
        merged_text = body.user_input
        accumulated = {}
        if body.accumulated_input:
            accumulated = extract_constraints(body.accumulated_input)
            merged_text = f"{body.accumulated_input}; {body.user_input}"

        constraints = extract_constraints(body.user_input)
        if accumulated:
            constraints = merge_constraints(accumulated, constraints)

        is_complete, missing_p0, missing_p1 = check_completeness(constraints)

        if is_complete and accumulated:
            # 合并后完整了 → 升格为 selection
            result["intent"] = "selection"
            result["merged_input"] = merged_text  # 传给前端用于选型
        else:
            response = build_clarification_response(merged_text, constraints)
            result["clarify_response"] = response
            result["missing_p0"] = missing_p0
            result["missing_p1"] = missing_p1

    return result


# ── 流式对话端点（非选型交互用）──────────────────────────────

@app.post("/agent/chat/stream")
async def agent_chat_stream_endpoint(body: AgentRequest):
    """轻量流式对话 — 不触发选型流水线，仅 ReAct Agent 自然语言交互。"""

    async def _stream_chat() -> AsyncGenerator[str, None]:
        import time as _time
        t0 = _time.time()
        try:
            from .react_agent import run_agent
            from .intent_classifier import classify
            import json as _json

            yield f"event: start\ndata: {_json.dumps({'status': 'agent_thinking'})}\n\n"

            result = run_agent(body.user_input, session_id=body.session_id,
                              thinking_depth=body.thinking_depth)

            # ── 思考流：暴露 ReAct 工具调用链 ────────────────────
            if body.thinking_depth != "off":
                tool_calls = result.get("tool_calls", [])
                for tc in tool_calls:
                    tool_name = tc.get("tool", "unknown")
                    tool_args = tc.get("args", {})
                    args_summary = ", ".join(
                        f"{k}={str(v)[:60]}" for k, v in tool_args.items()
                    )
                    yield f"event: thinking_delta\ndata: {_json.dumps({'stage': 'agent', 'text': f'调用工具：{tool_name}({args_summary})'}, ensure_ascii=False)}\n\n"
                    raw_result = tc.get("result", "")
                    if raw_result and body.thinking_depth in ("contemplation", "exhaustive"):
                        preview = str(raw_result)[:120].replace("\n", " ")
                        yield f"event: thinking_delta\ndata: {_json.dumps({'stage': 'agent', 'text': f'工具返回：{preview}'}, ensure_ascii=False)}\n\n"

            # 提取文本回复
            text = ""
            if isinstance(result, dict):
                text = result.get("response") or result.get("output") or str(result)
            else:
                text = str(result)

            yield f"event: text_delta\ndata: {_json.dumps({'text': text})}\n\n"

            # ── Memory：提取并记录用户称呼 ────────────────────
            try:
                from .memory import update_user_name, get_user_context
                import re as _re
                name_m = _re.search(r'(?:我是|我叫|叫我|name is|I am|I\'m)\s*[\x80-\uffff\w]+', body.user_input)
                if name_m:
                    raw = name_m.group(0)
                    name = _re.sub(r'(?:我是|我叫|叫我|name is|I am|I\'m)\s*', '', raw).strip().rstrip('.,;!。，；！')
                    if name and len(name) <= 10:
                        update_user_name(name)
            except Exception as e:
                import logging
                logging.getLogger(__name__).warning(f"agent_chat_stream: 记忆写入失败: {e}")

            elapsed = round(_time.time() - t0, 2)
            yield f"event: done\ndata: {_json.dumps({'elapsed_s': elapsed, 'intent': 'chat'})}\n\n"
        except Exception as e:
            yield f"event: error\ndata: {_json.dumps({'message': str(e)})}\n\n"

    return StreamingResponse(
        _stream_chat(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ── SSE 流式输出端点（B1 任务）──────────────────────────────────────

def _build_interaction_prompt(report) -> str:
    """选型完成后生成交互引导文案。"""
    recs = report.recommended_parts or []
    if not recs:
        return "\n\n> ⚠ 未找到满足条件的推荐器件，建议放宽约束后重试。"

    top1 = recs[0]
    pn = top1.part.part_number if hasattr(top1.part, "part_number") else "?"
    mfr = top1.part.manufacturer if hasattr(top1.part, "manufacturer") else ""
    score = int(top1.score.total_score) if hasattr(top1.score, "total_score") else 0
    risk = report.risks.overall_risk_level if hasattr(report.risks, "overall_risk_level") else "?"
    total_count = len(recs) + len([
        s for s in (report.candidates or [])
        if getattr(s, "recommendation_level", "") == "backup"
    ])

    return f"""
---

### 下一步操作

> 首选推荐：**#{1} {pn}**（{mfr}），综合评分 **{score} 分**，整体风险 **{risk.upper()}**

共 **{total_count}** 条候选器件（编号见上方列表）。请选择：

- **选择器件** — 输入编号（如 `1`）确认首选，系统将生成该器件的完整报告（BOM + 风险评估）
- **调整需求** — 输入修改后的参数（如 `输出电流改为 5A`、`换国产器件`）重新选型
- **导出 BOM** — 输入 `/export` 下载完整 BOM Excel 清单
- **查看报告** — 点击下方「查看报告」按钮浏览 BOM / 风险评估报告

> 💡 也可直接描述你的想法，如 _"就选第 1 款"_、_"有没有国产替代"_、_"帮我对比 #2 和 #3"_"""


def _rebuild_summary_from_cached_dict(cached_report: dict) -> str:
    """从缓存字典重建完整候选列表摘要（代替旧格式的 summary_markdown）。"""
    candidates = cached_report.get("candidates", [])
    user_input = cached_report.get("user_input", "")
    recommended = [c for c in candidates
                   if (c.get("recommendation_level") if isinstance(c, dict) else None) == "recommended"]
    backup = [c for c in candidates
              if (c.get("recommendation_level") if isinstance(c, dict) else None) == "backup"]

    lines = [
        "## 选型报告（缓存命中）",
        "",
        f"**需求**：{user_input}",
        "",
        f"**检索结果**：共 {len(candidates)} 条候选，"
        f"**{len(recommended)} 条推荐**，{len(backup)} 条备选。",
    ]

    def _fmt_part(idx: int, c: dict) -> list[str]:
        p = c.get("part", {}) if isinstance(c, dict) else {}
        s = c.get("score", {}) if isinstance(c, dict) else {}
        pn  = p.get("part_number", "?") if isinstance(p, dict) else "?"
        mfr = p.get("manufacturer", "—") if isinstance(p, dict) else "—"
        pkg = p.get("package", "") if isinstance(p, dict) else ""
        pkg_str = f" / {pkg}" if pkg else ""
        dom = " 🇨🇳" if (p.get("is_domestic") if isinstance(p, dict) else False) else ""
        total = int(s.get("total_score", 0)) if isinstance(s, dict) else 0
        param = int(s.get("parameter_match_score", 0)) if isinstance(s, dict) else 0
        supply = int(s.get("supply_risk_score", 0)) if isinstance(s, dict) else 0
        cost  = int(s.get("cost_score", 0)) if isinstance(s, dict) else 0
        dom_sc = int(s.get("domestic_score", 0)) if isinstance(s, dict) else 0
        out = [
            f"**#{idx}** `{pn}`{dom}  —  {mfr}{pkg_str}",
            f"- 综合得分：**{total}**（参数 {param} | 供应 {supply} | 成本 {cost} | 国产 {dom_sc}）",
        ]
        reasons = s.get("reasons", []) if isinstance(s, dict) else []
        good = [r for r in reasons if "✓" in r or "满足" in r][:2]
        for r in good:
            out.append(f"  - {r}")
        out.append("")
        return out

    if recommended:
        lines += ["", f"### 推荐器件（{len(recommended)} 条）", ""]
        for i, c in enumerate(recommended, start=1):
            lines.extend(_fmt_part(i, c))

    if backup:
        lines += [f"### 备选器件（{len(backup)} 条）", ""]
        start = len(recommended) + 1
        for i, c in enumerate(backup, start=start):
            p = c.get("part", {}) if isinstance(c, dict) else {}
            s = c.get("score", {}) if isinstance(c, dict) else {}
            pn  = p.get("part_number", "?") if isinstance(p, dict) else "?"
            mfr = p.get("manufacturer", "—") if isinstance(p, dict) else "—"
            total = int(s.get("total_score", 0)) if isinstance(s, dict) else 0
            param = int(s.get("parameter_match_score", 0)) if isinstance(s, dict) else 0
            supply = int(s.get("supply_risk_score", 0)) if isinstance(s, dict) else 0
            dom = " 🇨🇳" if (p.get("is_domestic") if isinstance(p, dict) else False) else ""
            lines.append(
                f"**#{i}** `{pn}`{dom}  —  {mfr}"
                f"  综合得分 **{total}**（参数 {param} | 供应 {supply}）"
            )
        lines.append("")

    risks = cached_report.get("risks", {})
    if isinstance(risks, dict):
        lvl = risks.get("overall_risk_level", "?")
        supply_sum = risks.get("supply_risk_summary", "")
        eng_sum = risks.get("engineering_risk_summary", "")
        risk_emoji = {"high": "⚠", "medium": "△", "low": "✓"}.get(lvl, "?")
        lines += [
            f"**整体风险**：{risk_emoji} {lvl.upper()}",
            f"- 供应链：{supply_sum}",
            f"- 工程：{eng_sum}",
        ]

    return "\n".join(lines)


def _safe_serialize(obj):
    """安全序列化 Pydantic/dataclass 对象为 JSON 兼容 dict。"""
    if hasattr(obj, 'dict'):
        return _safe_serialize(obj.dict())
    if hasattr(obj, 'model_dump'):
        return _safe_serialize(obj.model_dump())
    if isinstance(obj, dict):
        return {k: _safe_serialize(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_safe_serialize(v) for v in obj]
    return obj


def _yield_sse(event: str, data: dict) -> str:
    """构造 SSE 事件字符串（模块级，供各阶段子函数复用）。"""
    return f"event: {event}\ndata: {json.dumps(_safe_serialize(data), ensure_ascii=False)}\n\n"


# ═══════════════════════════════════════════════════════════════════
# 阶段子函数（拆分自 _stream_analyze，降低圈复杂度）
# 每个函数返回 (结果, sse_events_list)
# ═══════════════════════════════════════════════════════════════════

async def _stream_stage1_parse(req: AnalyzeRequest, loop) -> tuple:
    """Stage 1：需求解析 → (requirement, events)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "parse", "status": "started"}))
    from .requirement_parser import parse_requirement
    requirement = await loop.run_in_executor(None, parse_requirement, req.user_input)
    events.append(_yield_sse("parse_done", {
        "status": "需求解析完成",
        "constraints": _safe_serialize(requirement),
        "fields_parsed": sum(1 for v in [
            requirement.category, requirement.topology,
            requirement.input_voltage_nominal_v, requirement.output_voltage_v,
            requirement.output_current_a, requirement.temperature_min_c,
            requirement.temperature_max_c, requirement.grade,
        ] if v is not None),
    }))
    if req.thinking_depth != "off":
        think_parts = []
        if requirement.topology:
            think_parts.append(f"拓扑：{requirement.topology}")
        if requirement.input_voltage_nominal_v:
            think_parts.append(f"Vin={requirement.input_voltage_nominal_v}V")
        if requirement.output_voltage_v:
            think_parts.append(f"Vout={requirement.output_voltage_v}V")
        if requirement.output_current_a:
            think_parts.append(f"Iout={requirement.output_current_a}A")
        if requirement.temperature_min_c is not None and requirement.temperature_max_c is not None:
            think_parts.append(f"温度范围 {requirement.temperature_min_c}~{requirement.temperature_max_c}°C")
        if requirement.grade:
            think_parts.append(f"等级：{requirement.grade}")
        events.append(_yield_sse("thinking_delta", {
            "stage": "parse",
            "text": "约束提取：" + ("、".join(think_parts) if think_parts else "参数不完整，将进行宽松搜索"),
        }))
    return requirement, events


async def _stream_stage2_search(req: AnalyzeRequest, requirement, loop) -> tuple:
    """Stage 2：器件搜索 + 详情富化 + 候选放宽 → (candidates, events)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "search", "status": "started"}))
    from .ezplm_client import search_parts, enrich_candidates_with_details
    candidates = await loop.run_in_executor(None, search_parts, requirement)
    events.append(_yield_sse("search_done", {
        "status": "搜索完成",
        "candidate_count": len(candidates),
        "sources": list(set(getattr(c, "source", "unknown") for c in candidates)),
    }))
    if candidates:
        candidates = await loop.run_in_executor(
            None,
            lambda: enrich_candidates_with_details(candidates, max_enrich=6),
        )
    if req.thinking_depth != "off":
        if candidates:
            mfrs = list(dict.fromkeys(
                getattr(c, "manufacturer", "") for c in candidates
                if getattr(c, "manufacturer", "")
            ))[:3]
            mfr_str = "、".join(mfrs) if mfrs else "多品牌"
            events.append(_yield_sse("thinking_delta", {
                "stage": "search",
                "text": f"检索到 {len(candidates)} 个候选，来自 {mfr_str} 等厂商，按参数匹配→供应链→成本→国产化率四维评分",
            }))
        else:
            events.append(_yield_sse("thinking_delta", {
                "stage": "search",
                "text": "未找到精确匹配器件，将放宽电流约束后重新搜索",
            }))
    if not candidates:
        events.append(_yield_sse("warning", {"message": "未找到匹配器件，请检查需求或扩充数据源"}))
    if len(candidates) < 3 and not getattr(requirement, 'output_current_a', None) is None:
        old_iout = requirement.output_current_a
        requirement.output_current_a = round(old_iout * 0.8, 2)
        events.append(_yield_sse("stage", {
            "stage": "search",
            "status": f"candidates<3, relaxing Iout {old_iout}A->{requirement.output_current_a}A",
        }))
        candidates = await loop.run_in_executor(None, search_parts, requirement)
        requirement.output_current_a = old_iout
    return candidates, events


async def _stream_stage3_score(req: AnalyzeRequest, requirement, candidates, loop) -> tuple:
    """Stage 3：评分计算 → (scored, events)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "score", "status": "started", "total": len(candidates)}))
    from .scoring import score_candidates
    scored = await loop.run_in_executor(None, score_candidates, requirement, candidates)
    for i, s in enumerate(scored):
        events.append(_yield_sse("score_update", {
            "status": f"评分完成: {s.part.part_number}",
            "index": i + 1, "total": len(scored),
            "part_number": s.part.part_number,
            "manufacturer": s.part.manufacturer,
            "total_score": s.score.total_score,
            "parameter_match_score": s.score.parameter_match_score,
            "recommendation_level": s.recommendation_level,
            "scoring_mode": s.score.scoring_mode,
        }))
    if req.thinking_depth != "off" and scored:
        top = scored[0]
        score_detail = (
            f"参数 {int(top.score.parameter_match_score)}"
            f"+供应 {int(top.score.supply_risk_score)}"
            f"+成本 {int(top.score.cost_score)}"
            f"+国产 {int(top.score.domestic_score)}"
            f"= {int(top.score.total_score)} 分"
        )
        events.append(_yield_sse("thinking_delta", {
            "stage": "score",
            "text": f"首选：{top.part.part_number}（{top.part.manufacturer or '—'}）{score_detail}",
        }))
        if req.thinking_depth in ("contemplation", "exhaustive") and top.score.reasons:
            reasons_str = "；".join(top.score.reasons[:3])
            events.append(_yield_sse("thinking_delta", {
                "stage": "score",
                "text": f"评分依据：{reasons_str}",
            }))
    return scored, events


async def _stream_stage4_evidence(req: AnalyzeRequest, scored, requirement, loop) -> tuple:
    """Stage 4：证据构建 → (evidence, events)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "evidence", "status": "started"}))
    from .evidence import build_evidence
    evidence = await loop.run_in_executor(None, build_evidence, scored, requirement)
    avg_conf = round(sum(e.confidence for e in evidence) / len(evidence), 3) if evidence else 0.0
    evidence_items = []
    for e in evidence:
        evidence_items.append({
            "part_number": e.part_number,
            "claim": e.claim,
            "evidence_type": e.evidence_type,
            "source_field": e.source_field,
            "confidence": e.confidence,
            "need_human_review": getattr(e, "need_human_review", False),
        })
    events.append(_yield_sse("evidence_done", {
        "status": "证据链构建完成",
        "evidence_count": len(evidence),
        "avg_confidence": avg_conf,
        "evidence_items": evidence_items,
    }))
    if req.thinking_depth != "off" and evidence:
        high_conf = sum(1 for e in evidence if e.confidence >= 0.8)
        need_review = sum(1 for e in evidence if getattr(e, "need_human_review", False))
        events.append(_yield_sse("thinking_delta", {
            "stage": "evidence",
            "text": (
                f"证据链：{len(evidence)} 条参数符合性记录"
                f"，高置信度 {high_conf} 条（≥80%）"
                + (f"，{need_review} 条需人工复查" if need_review else "")
            ),
        }))
    return evidence, events


async def _stream_stage5_risk(req: AnalyzeRequest, requirement, scored, loop) -> tuple:
    """Stage 5：风险评估 → (risks, events)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "risk", "status": "started"}))
    from .report_generator import _assess_risks
    risks = await loop.run_in_executor(None, _assess_risks, requirement, scored)
    events.append(_yield_sse("risk_done", {
        "status": "风险评估完成",
        "overall_risk_level": risks.overall_risk_level,
        "risk_count": len(risks.risk_items),
        "high": sum(1 for r in risks.risk_items if r.severity == "high"),
        "medium": sum(1 for r in risks.risk_items if r.severity == "medium"),
        "low": sum(1 for r in risks.risk_items if r.severity == "low"),
        "supply_summary": risks.supply_risk_summary,
        "engineering_summary": risks.engineering_risk_summary,
    }))
    if req.thinking_depth != "off":
        high_c = sum(1 for r in risks.risk_items if r.severity == "high")
        med_c  = sum(1 for r in risks.risk_items if r.severity == "medium")
        low_c  = sum(1 for r in risks.risk_items if r.severity == "low")
        events.append(_yield_sse("thinking_delta", {
            "stage": "risk",
            "text": f"风险等级：{risks.overall_risk_level.upper()}（高 {high_c} / 中 {med_c} / 低 {low_c} 项）",
        }))
        if req.thinking_depth in ("contemplation", "exhaustive"):
            if risks.supply_risk_summary:
                events.append(_yield_sse("thinking_delta", {
                    "stage": "risk",
                    "text": f"供应链：{risks.supply_risk_summary}",
                }))
            if risks.engineering_risk_summary:
                events.append(_yield_sse("thinking_delta", {
                    "stage": "risk",
                    "text": f"工程设计：{risks.engineering_risk_summary}",
                }))
    return risks, events


async def _stream_stage6_report(req: AnalyzeRequest, requirement, scored, evidence, risks,
                                 t_start: float, loop) -> tuple:
    """Stage 6+7：报告生成 + 完成事件 → (events, report)"""
    events: list = []
    events.append(_yield_sse("stage", {"stage": "report", "status": "started"}))
    from .report_generator import build_report
    report = await loop.run_in_executor(None, build_report, requirement, scored, evidence)
    sid = req.session_id or _DEFAULT_SESSION_ID
    _session_reports[sid] = report
    _session_constraints[sid] = requirement
    if report.summary_markdown:
        for line in report.summary_markdown.split('\n'):
            if line.strip():
                events.append(_yield_sse("text_delta", {"text": line}))
    interaction_text = _build_interaction_prompt(report)
    for line in interaction_text.split('\n'):
        if line.strip():
            events.append(_yield_sse("text_delta", {"text": line}))
    elapsed = round(time.time() - t_start, 2)
    events.append(_yield_sse("done", {
        "status": "分析完成",
        "elapsed_s": elapsed,
        "request_id": report.request_id,
        "recommended_count": len(report.recommended_parts),
        "candidate_count": len(report.candidates),
        "overall_risk": risks.overall_risk_level,
        "summary": report.summary_markdown or "",
    }))
    return events, report


async def _stream_analyze(req: AnalyzeRequest) -> AsyncGenerator[str, None]:
    """异步生成器：按阶段推送 SSE 事件（含 B4 语义缓存集成）。

    各阶段已拆分为独立子函数：parse → search → score → evidence → risk → report。
    """
    t_start = time.time()
    loop = asyncio.get_running_loop()

    try:
        # ── B4：语义缓存层检查 ──────────────────────────────────
        from .semantic_cache import get_semantic_cache
        cache = get_semantic_cache()
        cache_result = cache.get(req.user_input)

        if cache_result is not None:
            elapsed = round(time.time() - t_start, 2)
            yield _yield_sse("cache_hit", {"hit": True, "similarity": cache_result.get("similarity", 0)})
            cached_report = cache_result.get("cached_result", {})

            scored = cached_report.get("scored_parts", []) or cached_report.get("candidates", [])
            for i, s in enumerate(scored):
                part = s.get("part", {}) if isinstance(s, dict) else getattr(s, "part", None)
                score = s.get("score", {}) if isinstance(s, dict) else getattr(s, "score", None)
                if part:
                    pn = part.get("part_number", "") if isinstance(part, dict) else getattr(part, "part_number", "")
                    total_score = score.get("total_score", 0) if isinstance(score, dict) else (getattr(score, "total_score", 0) if score else 0)
                    rec_level = s.get("recommendation_level", "") if isinstance(s, dict) else getattr(s, "recommendation_level", "")
                    yield _yield_sse("score_update", {
                        "status": f"缓存命中: {pn}",
                        "index": i + 1, "total": len(scored),
                        "part_number": pn,
                        "total_score": total_score,
                        "recommendation_level": rec_level,
                    })

            summary = _rebuild_summary_from_cached_dict(cached_report)
            if summary:
                for line in summary.split('\n'):
                    if line.strip():
                        yield _yield_sse("text_delta", {"text": line})

            recs = cached_report.get("recommended_parts", [])
            all_candidates = cached_report.get("candidates", [])
            total_count = len(all_candidates)
            if recs:
                r0 = recs[0]
                part = r0.get("part", {}) if isinstance(r0, dict) else {}
                score = r0.get("score", {}) if isinstance(r0, dict) else {}
                pn = part.get("part_number", "?") if isinstance(part, dict) else "?"
                mfr = part.get("manufacturer", "") if isinstance(part, dict) else ""
                total_score = int(score.get("total_score", 0)) if isinstance(score, dict) else 0
                risks = cached_report.get("risks", {})
                risk_level = (risks.get("overall_risk_level", "?") if isinstance(risks, dict) else "?").upper()
                interaction_text = f"""
---

### 下一步操作

> 首选推荐：**#1 {pn}**（{mfr}），综合评分 **{total_score} 分**，整体风险 **{risk_level}**

共 **{total_count}** 条候选器件（编号见上方列表）。请选择：

- **选择器件** — 输入编号（如 `1`）确认首选，系统将生成该器件的完整报告（BOM + 风险评估）
- **调整需求** — 输入修改后的参数（如 `输出电流改为 5A`、`换国产器件`）重新选型
- **导出 BOM** — 输入 `/export` 下载完整 BOM Excel 清单
- **查看报告** — 点击下方「查看报告」按钮浏览 BOM / 风险评估报告

> 💡 也可直接描述你的想法，如 _"就选第 1 款"_、_"有没有国产替代"_、_"帮我对比 #2 和 #3"_
"""
                for line in interaction_text.split('\n'):
                    if line.strip():
                        yield _yield_sse("text_delta", {"text": line})

            rec_count = len(cached_report.get("recommended_parts", []))
            risk_val = cached_report.get("risks", {})
            if isinstance(risk_val, dict):
                overall = risk_val.get("overall_risk_level", "?")
            else:
                overall = getattr(risk_val, "overall_risk_level", "?")

            yield _yield_sse("done", {
                "status": "分析完成（语义缓存命中）",
                "elapsed_s": elapsed,
                "request_id": cached_report.get("request_id", "cached"),
                "recommended_count": rec_count,
                "candidate_count": len(cached_report.get("candidates", [])),
                "overall_risk": overall,
                "summary": summary,
                "cache_hit": True,
            })
            return

        yield _yield_sse("cache_hit", {"hit": False})

        # ── Stage 1：需求解析 ──────────────────────────────────
        requirement, events = await _stream_stage1_parse(req, loop)
        for ev in events:
            yield ev

        # ── Stage 2：器件搜索 ──────────────────────────────────
        candidates, events = await _stream_stage2_search(req, requirement, loop)
        for ev in events:
            yield ev

        # ── Stage 3：评分计算 ──────────────────────────────────
        scored, events = await _stream_stage3_score(req, requirement, candidates, loop)
        for ev in events:
            yield ev

        # ── Stage 4：证据构建 ──────────────────────────────────
        evidence, events = await _stream_stage4_evidence(req, scored, requirement, loop)
        for ev in events:
            yield ev

        # ── P1/P5: CriticNode 自省检查 ─────────────────────
        from .langgraph_orchestrator import critic_node
        critic_state = {"constraints": requirement, "scored": scored, "retry_count": 0}
        critic_result = critic_node(critic_state)
        if not critic_result.get("critic_passed"):
            yield _yield_sse("warning", {"message": f"Critic: {critic_result.get('error', '')}"})

        # ── Stage 5：风险评估 ──────────────────────────────────
        risks, events = await _stream_stage5_risk(req, requirement, scored, loop)
        for ev in events:
            yield ev

        # ── Stage 6：报告生成 + 完成 ───────────────────────────
        events, report = await _stream_stage6_report(req, requirement, scored, evidence, risks, t_start, loop)
        for ev in events:
            yield ev

        # ── B4：将结果存入语义缓存 ────────────────────────────
        try:
            from .semantic_cache import get_semantic_cache
            get_semantic_cache().set(req.user_input, report.dict())
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"语义缓存写入失败: {e}")

        # ── Memory：记录选型历史 ────────────────────────────────
        try:
            from .memory import record_selection
            recs = report.recommended_parts
            summary_text = f"推荐 {len(recs)} 款" + (f", Top1={recs[0].part.part_number}" if recs else ", 无推荐")
            record_selection(req.user_input, summary_text)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning(f"选型历史记录失败: {e}")

    except Exception as exc:
        elapsed = round(time.time() - t_start, 2)
        yield _yield_sse("error", {
            "status": "错误",
            "message": str(exc),
            "elapsed_s": elapsed,
            "traceback": traceback.format_exc()[:500],
        })


@app.post("/analyze/stream")
async def analyze_stream_endpoint(body: AnalyzeRequest):
    """流式输出端点：SSE 逐段推送选型报告

    ── B4：语义缓存支持 ──
    - 响应头 X-Cache: HIT 表示命中缓存
    - 响应头 X-Cache: MISS 表示未命中缓存

    示例事件流：
    - cache_hit: 缓存命中状态（HIT/MISS）
    - parse_done: 需求解析完成
    - search_done: 搜索完成 + 候选数量
    - score_update: 评分完成（多次）
    - evidence_done: 证据链完成
    - risk_done: 风险评估完成 + RiskIR
    - text_delta: 报告文本片段（多次）
    - done: 完成 + 总耗时 + 完整报告
    """
    # ── B4：检查语义缓存 ──────────────────────────────────────
    from .semantic_cache import get_semantic_cache
    cache = get_semantic_cache()
    cache_result = cache.get(body.user_input)

    # 根据缓存状态设置响应头
    cache_header = "HIT" if cache_result is not None else "MISS"

    return StreamingResponse(
        _stream_analyze(body),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
            "X-Cache": cache_header,
        }
    )


# ── 电路图生成端点（B2 任务）────────────────────────────────────────

@app.get("/schematic/{topology}")
async def get_schematic(topology: str, Vin: float, Vout: float, Iout: float):
    """生成参数化应用电路 SVG

    Args:
        topology: 拓扑类型 ('buck', 'boost', 'ldo')
        Vin: 输入电压 (V)
        Vout: 输出电压 (V)
        Iout: 输出电流 (A)

    Returns:
        SVG 格式的电路图
    """
    try:
        from .schematic_generator import generate_schematic
        svg = generate_schematic(topology, Vin, Vout, Iout)
        return Response(content=svg, media_type="image/svg+xml")
    except ValueError as e:
        return JSONResponse(
            status_code=400,
            content={"detail": str(e)},
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"detail": f"电路图生成错误: {str(e)}"},
        )


# ── 三类报告 Markdown 端点 ───────────────────────────────────────

@app.get("/report/{report_type}")
async def get_report(report_type: str, session_id: Optional[str] = None):
    """返回指定会话的三类报告 Markdown 内容。

    Args:
        report_type: 'bom' | 'risk' | 'topology'
        session_id: 可选会话 ID（未提供时使用默认会话）

    Returns:
        {"content": "Markdown文本", "type": "bom|risk|topology"}
    """
    sid = session_id or _DEFAULT_SESSION_ID
    _latest_report = _session_reports.get(sid)
    _latest_constraints = _session_constraints.get(sid)
    if _latest_report is None:
        return JSONResponse(status_code=404, content={"detail": "暂无分析报告，请先执行一次选型分析"})

    try:
        from .output_generator import generate_all_reports
        from .output_bom import generate_bom
        from .output_generator import generate_risk_report, generate_topology
        import tempfile, os as _os

        rag_context = ""
        try:
            from .output_generator import _extract_rag_context
            rag_context = _extract_rag_context(_latest_report)
        except Exception:
            pass

        if report_type == "bom":
            md = generate_bom(_latest_report, rag_context=rag_context)
        elif report_type == "risk":
            md = generate_risk_report(_latest_report, _latest_constraints, rag_context=rag_context)
        elif report_type == "topology":
            md = generate_topology(_latest_constraints, _latest_report, rag_context=rag_context)
        else:
            return JSONResponse(status_code=400, content={"detail": f"未知报告类型: {report_type}，支持 bom/risk/topology"})

        return {"content": md, "type": report_type}
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"报告生成错误: {str(e)}"})


# ── 文件上传/解析端点 ──────────────────────────────────────────

ALLOWED_EXTENSIONS = {".pdf", ".csv", ".txt", ".md", ".json", ".xlsx", ".xls"}

@app.post("/upload/parse")
async def upload_and_parse(file: UploadFile = File(...)):
    """上传文件并提取文本内容（供 LLM 解析）。

    支持: PDF (数据手册), CSV (BOM表), TXT/MD, JSON, Excel
    """
    if not file.filename:
        return JSONResponse(status_code=400, content={"detail": "未提供文件"})

    ext = file.filename.lower().rsplit(".", 1)[-1] if "." in file.filename else ""
    if f".{ext}" not in ALLOWED_EXTENSIONS:
        return JSONResponse(status_code=400,
            content={"detail": f"不支持的文件类型: .{ext}，支持: {', '.join(ALLOWED_EXTENSIONS)}"})

    try:
        content_bytes = await file.read()

        if ext in ("txt", "md", "json", "csv"):
            content = content_bytes.decode("utf-8", errors="replace")
        elif ext == "pdf":
            try:
                import io
                from PyPDF2 import PdfReader
                reader = PdfReader(io.BytesIO(content_bytes))
                pages = [p.extract_text() or "" for p in reader.pages[:30]]  # 前30页（数据手册通常40-100页）
                content = "\n\n".join(pages)
            except ImportError:
                content = f"[PDF文件: {file.filename}] — PyPDF2 未安装，无法提取文本"
        elif ext in ("xlsx", "xls"):
            try:
                import io
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True)
                sheets = []
                total_rows = 0
                max_total_rows = 10000  # 总行数上限，防止内存溢出
                for name in wb.sheetnames:  # 读取所有 Sheet
                    ws = wb[name]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        if total_rows >= max_total_rows:
                            break
                        rows.append("\t".join(str(c) if c else "" for c in row))
                        total_rows += 1
                    sheets.append(f"=== {name} ===\n" + "\n".join(rows))
                    if total_rows >= max_total_rows:
                        break
                content = "\n\n".join(sheets)
            except ImportError:
                content = f"[Excel文件: {file.filename}] — openpyxl 未安装"
        else:
            content = f"[不支持的文件类型: {file.filename}]"

        return {
            "filename": file.filename,
            "type": ext,
            "size_bytes": len(content_bytes),
            "content": content[:5000],  # 限制返回大小
            "truncated": len(content) > 5000,
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"文件解析错误: {str(e)}"})


# ── BOM Excel 导出端点（B6）────────────────────────────────────

@app.post("/export/bom")
async def export_bom_endpoint(session_id: Optional[str] = None):
    """导出工程级 BOM Excel 文件（三 Sheet）。

    Args:
        session_id: 可选会话 ID（未提供时使用默认会话）

    Returns:
        .xlsx binary (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)
    """
    sid = session_id or _DEFAULT_SESSION_ID
    _latest_report = _session_reports.get(sid)
    if _latest_report is None:
        return JSONResponse(status_code=404, content={"detail": "暂无分析报告，请先执行一次选型分析"})

    try:
        from .output_bom import generate_bom_excel
        xlsx_bytes = generate_bom_excel(_latest_report)
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=BOM_{_latest_report.request_id[:8]}.xlsx",
            },
        )
    except Exception as e:
        return JSONResponse(status_code=500, content={"detail": f"Excel 生成错误: {str(e)}"})
